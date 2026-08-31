#!/usr/bin/env python3
"""
build_seed.py — regenerate Postgres seed CSVs from the two source Excel files.

Sources:
  1. albero_tabellare.xlsx (sheet 'ALBERO ECR')
     LIVELLO 1, COD.LIV1, LIVELLO 2, COD.LIV2, LIVELLO 3, COD.LIV3,
     LIVELLO 4, COD.LIV4, UdM, LIVELLO 5, COD.LIV5, cod.completo

  2. New_Report.xlsx (sheet 'Blank Report')
     Reparto, Gruppo, Settore, Famiglia, Prodotto, <unnamed description column>
     Row 2 is a junk placeholder row ('REPARTO???' ...) and is skipped.

Outputs (written to server/seed/ next to this script, unless -o/--out-dir given):
  - ecr_albero.csv
  - catalogo_prodotti.csv
  - famiglie.csv

Usage:
  python3 build_seed.py [albero_xlsx] [report_xlsx] [-o OUT_DIR]

Both positional arguments are optional; they default to the paths this
script was originally built against (see DEFAULT_ALBERO_PATH /
DEFAULT_REPORT_PATH below). Re-run this script whenever the source
spreadsheets are updated.
"""

import argparse
import csv
import os
import re
import sys

import openpyxl

DEFAULT_ALBERO_PATH = "/tmp/albero/albero_tabellare.xlsx"
DEFAULT_REPORT_PATH = "/root/.claude/uploads/bd75985c-a07c-54c2-ad32-b303290272e9/a7d16eab-New_Report.xlsx"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

ALBERO_SHEET = "ALBERO ECR"
REPORT_SHEET = "Blank Report"


def normalize_descrizione(s):
    """
    Normalize a product description for matching:
      - uppercase
      - runs of '*' -> single space
      - runs of '.' that sit at end-of-string or right before whitespace -> single space
      - collapse whitespace runs to a single space
      - strip leading/trailing whitespace

    Example: 'SFOGLIE ZUCC.200 **MONTAG***.' -> 'SFOGLIE ZUCC.200 MONTAG'
    """
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"\*+", " ", s)
    s = re.sub(r"\.+(?=\s|$)", " ", s)
    s = re.sub(r"\s+", " ", s)
    s = s.strip().upper()
    return s


def build_ecr_albero(albero_path, out_path):
    wb = openpyxl.load_workbook(albero_path, read_only=True, data_only=True)
    ws = wb[ALBERO_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)  # noqa: F841 (header row, positions are known/fixed)

    out_header = [
        "cod_famiglia",
        "reparto",
        "cod_reparto",
        "gruppo",
        "cod_gruppo",
        "sottogruppo",
        "cod_sottogruppo",
        "famiglia",
        "cod_famiglia_liv4",
        "udm",
    ]

    n_written = 0
    n_skipped = 0
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(out_header)
        for row in rows_iter:
            if row is None:
                n_skipped += 1
                continue
            (
                liv1, cod_liv1,
                liv2, cod_liv2,
                liv3, cod_liv3,
                liv4, cod_liv4,
                udm,
                liv5, cod_liv5,
                cod_completo,
            ) = (list(row) + [None] * 12)[:12]

            if cod_completo is None or str(cod_completo).strip() == "":
                n_skipped += 1
                continue

            cod_completo = str(cod_completo).strip()
            cod_famiglia = cod_completo[:8]

            writer.writerow([
                cod_famiglia,
                liv1,
                cod_liv1,
                liv2,
                cod_liv2,
                liv3,
                cod_liv3,
                liv4,
                cod_liv4,
                udm,
            ])
            n_written += 1

    return n_written, n_skipped


def build_catalogo_prodotti(report_path, out_path):
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    ws = wb[REPORT_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)  # noqa: F841

    out_header = [
        "cod_prodotto",
        "descrizione",
        "reparto",
        "gruppo",
        "settore",
        "famiglia",
        "descrizione_norm",
    ]

    # famiglie aggregation: (reparto, gruppo, settore, famiglia) -> count
    famiglie_counts = {}

    n_written = 0
    n_skipped = 0

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(out_header)
        for row in rows_iter:
            if row is None:
                n_skipped += 1
                continue
            row = list(row) + [None] * (6 - len(row))
            reparto, gruppo, settore, famiglia, prodotto, descrizione = row[:6]

            # Skip the junk placeholder row and any row with empty Reparto.
            if reparto is None or str(reparto).strip() == "":
                n_skipped += 1
                continue
            if str(reparto).strip().upper() == "REPARTO???":
                n_skipped += 1
                continue

            descrizione_str = "" if descrizione is None else str(descrizione)
            descrizione_norm = normalize_descrizione(descrizione_str)

            writer.writerow([
                prodotto,
                descrizione_str,
                reparto,
                gruppo,
                settore,
                famiglia,
                descrizione_norm,
            ])
            n_written += 1

            key = (reparto, gruppo, settore, famiglia)
            famiglie_counts[key] = famiglie_counts.get(key, 0) + 1

    return n_written, n_skipped, famiglie_counts


def build_famiglie(famiglie_counts, out_path):
    out_header = ["reparto", "gruppo", "settore", "famiglia", "n_prodotti"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(out_header)
        for (reparto, gruppo, settore, famiglia), count in sorted(
            famiglie_counts.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "", kv[0][2] or "", kv[0][3] or "")
        ):
            writer.writerow([reparto, gruppo, settore, famiglia, count])
    return len(famiglie_counts)


def count_data_rows(path):
    with open(path, newline="", encoding="utf-8") as f:
        return sum(1 for _ in f) - 1  # minus header


def main():
    parser = argparse.ArgumentParser(description="Build Postgres seed CSVs from source Excel files.")
    parser.add_argument("albero_xlsx", nargs="?", default=DEFAULT_ALBERO_PATH,
                         help="Path to albero_tabellare.xlsx (sheet 'ALBERO ECR')")
    parser.add_argument("report_xlsx", nargs="?", default=DEFAULT_REPORT_PATH,
                         help="Path to New_Report.xlsx (sheet 'Blank Report')")
    parser.add_argument("-o", "--out-dir", default=SCRIPT_DIR,
                         help="Directory to write output CSVs into (default: this script's directory)")
    args = parser.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)

    ecr_path = os.path.join(args.out_dir, "ecr_albero.csv")
    catalogo_path = os.path.join(args.out_dir, "catalogo_prodotti.csv")
    famiglie_path = os.path.join(args.out_dir, "famiglie.csv")

    print(f"Reading albero source: {args.albero_xlsx}")
    ecr_written, ecr_skipped = build_ecr_albero(args.albero_xlsx, ecr_path)
    print(f"  -> {ecr_path} ({ecr_written} rows written, {ecr_skipped} skipped)")

    print(f"Reading report source: {args.report_xlsx}")
    cat_written, cat_skipped, famiglie_counts = build_catalogo_prodotti(args.report_xlsx, catalogo_path)
    print(f"  -> {catalogo_path} ({cat_written} rows written, {cat_skipped} skipped)")

    n_famiglie = build_famiglie(famiglie_counts, famiglie_path)
    print(f"  -> {famiglie_path} ({n_famiglie} distinct reparto/gruppo/settore/famiglia combos)")

    print()
    print("Summary:")
    for path in (ecr_path, catalogo_path, famiglie_path):
        n_rows = count_data_rows(path)
        size = os.path.getsize(path)
        print(f"  {os.path.relpath(path)}: {n_rows} data rows, {size:,} bytes")


if __name__ == "__main__":
    sys.exit(main())
